import csv
import os
from io import BytesIO
from flask import Flask, render_template, request, send_file, redirect, url_for, flash, make_response
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime
from weasyprint import HTML
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.utils import secure_filename
from dateutil.parser import parse as date_parse
from difflib import get_close_matches
from functools import wraps
from flask import abort
from dotenv import load_dotenv
from flask_babel import Babel, _, lazy_gettext as _l
from flask import session
import uuid
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer
import logging
from flask import send_from_directory
from flask_talisman import Talisman

logging.basicConfig(level=logging.INFO)

load_dotenv()
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "fallback-secret")
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv("SQLALCHEMY_DATABASE_URI")
app.config['BABEL_DEFAULT_LOCALE'] = 'en'
app.config['BABEL_TRANSLATION_DIRECTORIES'] = 'translations'

print("Loaded DB URI:", app.config['SQLALCHEMY_DATABASE_URI'])


app.config['SESSION_COOKIE_SECURE'] = False  # if you're on HTTP
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

UPLOAD_FOLDER = os.path.join(app.root_path, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['BABEL_SUPPORTED_LOCALES'] = ['en', 'ms', 'id', 'zh_Hans'] 
app.config.update(
    MAIL_SERVER='smtp.gmail.com', 
    MAIL_PORT=587,
    MAIL_USE_TLS=True,
    MAIL_USERNAME=os.getenv('MAIL_USERNAME'),
    MAIL_PASSWORD=os.getenv('MAIL_PASSWORD'),
    MAIL_DEFAULT_SENDER=os.getenv('MAIL_DEFAULT_SENDER')
)

csp = {
    'default-src': ["'self'"],
    'style-src': ["'self'", "'unsafe-inline'", 'https://fonts.googleapis.com'],
    'font-src': ["'self'", 'https://fonts.gstatic.com'],
    'script-src': ["'self'", "'unsafe-inline'"]
}

Talisman(app, content_security_policy=csp)

mail = Mail(app) 
babel = Babel(app) 
db = SQLAlchemy(app)
migrate = Migrate(app, db)

latest_summary = None  
latest_mode = "date"

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# database tables
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.Text, nullable=False)
    plan = db.Column(db.String(20), default="free")
    is_admin = db.Column(db.Boolean, default=False)


class Upload(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    mode = db.Column(db.String(20))
    total = db.Column(db.Float)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    user = db.relationship('User', backref='uploads')

class FileSummary(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    file_id = db.Column(db.Integer, db.ForeignKey('upload.id'), nullable=False)
    summary_text = db.Column(db.Text)
    generated_at = db.Column(db.DateTime, default=datetime.utcnow)

class PaymentRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount = db.Column(db.Numeric(10, 2), nullable=False)
    currency = db.Column(db.String(3), default='BND', nullable=False)
    proof_filename = db.Column(db.String(256), nullable=True)
    status = db.Column(db.String(20), default='pending')  # pending, approved, rejected
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


# admin credentials
def create_admin_user():
    if not User.query.filter_by(username='admin').first():
        admin_password = os.getenv("ADMIN_PASSWORD")
        if not admin_password:
            raise ValueError("ADMIN_PASSWORD environment variable not set.")

        admin = User(
            username="admin",
            email="admin@example.com",
            password_hash=generate_password_hash(admin_password),
            is_admin=True,
            plan="premium"
        )
        db.session.add(admin)
        db.session.commit()


# load users
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# user can pick any language
def get_locale():
    lang = session.get('lang')
    logging.info(f"Current session lang: {lang}")
    logging.info(f"Locale selected: {lang}")
    return lang or request.accept_languages.best_match(['en', 'ms', 'id', 'zh_Hans'])

babel.locale_selector_func = get_locale
app.jinja_env.globals['get_locale'] = get_locale


# to change langauge
@app.route('/change_lang/<lang_code>')
def change_lang(lang_code):
    if lang_code not in app.config['BABEL_SUPPORTED_LOCALES']:
        flash("Unsupported language selected.")
        return redirect(request.referrer or url_for('landing'))
    
    session['lang'] = lang_code
    session.modified = True  # << ensure session writes change
    logging.info(f"Language changed to: {lang_code}")
    logging.info(f"User requested language change to: {lang_code}")
    return redirect(request.referrer or url_for('landing'))


# admin only no other user
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)  # forbidden
        return f(*args, **kwargs)
    return decorated_function
    
# to admin page
@app.route('/admin')
@login_required
@admin_required
def admin():
    users = User.query.all()
    uploads = Upload.query.all()  # if uploads exist
    payment_requests = PaymentRequest.query.order_by(PaymentRequest.created_at.desc()).all()
    return render_template('admin.html', users=users, uploads=uploads, payment_requests=payment_requests)

# delete a user from admin page
@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    if not current_user.is_admin:
        abort(403)

    user = User.query.get_or_404(user_id)

    if user.is_admin:
        flash("You cannot delete an admin user")
        return redirect(url_for('admin'))
    
    db.session.delete(user)
    db.session.commit()
    
    flash(f"User {user.username} has been deleted.", "success")
    return redirect(url_for('admin'))

# to signup page
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email') 
        password = request.form.get('password')

        if User.query.filter_by(username=username).first():
            flash("Username already exists")
            return redirect(url_for('signup'))

        hashed_password = generate_password_hash(password)
        new_user = User(username=username, email=email, password_hash=hashed_password)
        db.session.add(new_user)
        db.session.commit()

        flash("Signup successful! Please login.", "success")
        return redirect(url_for('login'))

    return render_template('signup.html')

# to login page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash("Logged in successfully", "success")
            if user.is_admin:
                return redirect(url_for('admin'))
            else:
                return redirect(url_for('index'))  # redirect to index after login
        else:
            flash("Invalid username or password")
            return redirect(url_for('login'))

    return render_template('login.html')

# to logout page
@app.route('/logout')
@login_required
def logout():
    session.pop('latest_summary', None)
    session.pop('latest_mode', None)
    logout_user()
    flash("Logged out successfully", "success")
    return redirect(url_for('landing'))


# user forget password
def generate_reset_token(user_id, expires_sec=3600):
    serializer = URLSafeTimedSerializer(app.secret_key)
    return serializer.dumps(user_id, salt='password-reset-salt')

def verify_reset_token(token, expires_sec=3600):
    serializer = URLSafeTimedSerializer(app.secret_key)
    try:
        user_id = serializer.loads(token, salt='password-reset-salt', max_age=expires_sec)
    except Exception:
        return None
    return User.query.get(user_id)

# to forget password page
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email_or_username = request.form.get('email_or_username')
        user = User.query.filter((User.username == email_or_username) | (User.email == email_or_username)).first()

        if user and user.email:
            token = generate_reset_token(user.id)
            reset_url = url_for('reset_password', token=token, _external=True)
            msg = Message("Password Reset Request", recipients=[user.email])
            msg.body = f"""To reset your password, visit this link:
{reset_url}

If you didn't request this, ignore this email.
"""
            mail.send(msg)
            flash('Password reset email sent! Check your inbox.')
        else:
            flash('No account found with that username or email.')
        return redirect(url_for('login'))

    return render_template('forgot_password.html')


# reset pwd
@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    user = verify_reset_token(token)
    if not user:
        flash('Invalid or expired token', 'danger')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        new_password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if new_password != confirm_password:
            flash("Passwords do not match", "danger")
            return redirect(request.url)

        if not new_password:
            flash('Please enter a new password', 'warning')
            return redirect(request.url)

        user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash('Your password has been updated. Please login.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html')



# load sales data

# any col name
DATE_HEADERS = ['date', 'Date', 'Date of Sale', 'sale_date']
ITEM_HEADERS = ['item', 'Item', 'product', 'Product Name']
AMOUNT_HEADERS = ['amount', 'Amount', 'total', 'Total Sales']


# auto dtect date
def parse_date_flexible(date_str):
    date_formats = [
        "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y",
        "%Y/%m/%d", "%Y.%m.%d", "%d %b %Y", "%d %B %Y"
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except ValueError:
            continue
            
    try:
        dt = date_parse(date_str, dayfirst=True)  # dayfirst=True bc dd/mm/yyyy
        return dt.date()
    except Exception:
        raise ValueError(f"Date '{date_str}' is not in a recognized format")

# auto detection for header
def detect_column(headers, candidates):
    headers_clean = [h.strip().lower() for h in headers]
    candidates_clean = [c.strip().lower() for c in candidates]
    
    for candidate in candidates_clean:
        match = get_close_matches(candidate, headers_clean, n=1, cutoff=0.8)
        if match:
            index = headers_clean.index(match[0])
            return headers[index]
    return None

# menganu summary glblity
def serialize_summary(summary, mode):
    if mode == "date":
        return {key.strftime("%d-%m-%Y"): val for key, val in summary.items()}
    elif mode == "combined":
        return {f"{key[0].strftime('%d-%m-%Y')}|{key[1]}": val for key, val in summary.items()}
    else:  # item
        return summary

def deserialize_summary(serialized, mode):
    if mode == "date":
        return {datetime.strptime(k, "%d-%m-%Y").date(): v for k, v in serialized.items()}
    elif mode == "combined":
        result = {}
        for k, v in serialized.items():
            date_str, item = k.split("|", 1)
            result[(datetime.strptime(date_str, "%d-%m-%Y").date(), item)] = v
        return result
    else:
        return serialized
    
 # meganu2
def get_current_summary():
    serialized = session.get("latest_summary")
    mode = session.get("latest_mode", "date")
    if not serialized:
        return None, mode
    return deserialize_summary(serialized, mode), mode


# loader sales data1
def load_sales_data(file_stream, mode="date", from_date=None, to_date=None):
    sales = defaultdict(float)

    # detect delimiter
    def detect_delimiter(sample):
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(sample, delimiters=[',', ';', '\t'])
            return dialect.delimiter
        except csv.Error:
            return ','
        
    file_stream.seek(0)
    sample = file_stream.read(2048).decode('utf-8')
    delimiter = detect_delimiter(sample)

    file_stream.seek(0)
    lines = file_stream.read().decode('utf-8').splitlines()
    reader = csv.DictReader(lines, delimiter=delimiter)


    # auto detect columns
    headers = reader.fieldnames
    date_col = detect_column(headers, DATE_HEADERS)
    item_col = detect_column(headers, ITEM_HEADERS)
    amount_col = detect_column(headers, AMOUNT_HEADERS)

    # identify extra column
    required_cols = {date_col, item_col, amount_col}
    extra_cols = [h for h in headers if h not in required_cols]

    if extra_cols:
        flash(f"Ignoring extra columns: {', '.join(extra_cols)}")

    missing = []
    if not date_col:
        missing.append("Date")
    if not item_col:
        missing.append("Item")
    if not amount_col:
        missing.append("Amount")

    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
            
    # if not date_col or not item_col or not amount_col:
    #     raise ValueError("CSV is missing one or more required columns")

    skipped_rows = 0
    for row in reader:
        try:
            amount = float(row[amount_col])
            date_obj = parse_date_flexible(row[date_col].strip())
            item = row[item_col].strip()

            if from_date and date_obj < from_date:
                continue
            if to_date and date_obj > to_date:
                continue

            if mode == "date":
                sales[date_obj] += amount
            elif mode == "item":
                sales[item] += amount
            elif mode == "combined":
                sales[(date_obj, item)] += amount
        except (ValueError, KeyError):
            skipped_rows +=1
            continue

    if skipped_rows:
        flash(f"Skipped {skipped_rows} invalid rows during import")

    if mode == "date":
        return dict(sorted(sales.items()))
    elif mode == "item":
        return dict(sorted(sales.items(), key=lambda x: x[0].lower()))
    elif mode == "combined":
        return dict(sorted(sales.items(), key=lambda x: (x[0][0], x[0][1].lower())))
    
# generate excel report
def generate_excel_report(summary, mode="date"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Summary"

    # set headers
    if mode == "date":
        ws.append(["Date", "Total Sales ($)"])
    elif mode == "combined":
        ws.append(["Date", "Item", "Total Sales ($)"])
    else:
        ws.append(["Item", "Total Sales ($)"])

    # style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # data rows
    for key, total in summary.items():
        if mode == "date":
            row = [key.strftime("%d/%m/%Y"), total]
        elif mode == "combined":
            row = [key[0].strftime("%d/%m/%Y"), key[1], total]
        else:
            row = [key, total]

        ws.append(row)
        # format amount as currency
        ws.cell(row=ws.max_row, column=len(row)).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=ws.max_row, column=len(row)).alignment = Alignment(horizontal="center")

    # total row
    total_sales = sum(summary.values())
    total_row_index = ws.max_row + 1

    if mode == "combined":
        ws.append(["", "Total", total_sales])
        ws.cell(row=total_row_index, column=2).font = Font(bold=True)
        ws.cell(row=total_row_index, column=2).alignment = Alignment(horizontal="center")

        ws.cell(row=total_row_index, column=3).font = Font(bold=True)
        ws.cell(row=total_row_index, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=total_row_index, column=3).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    else:
        ws.append(["Total", total_sales])
        ws.cell(row=total_row_index, column=1).font = Font(bold=True)
        ws.cell(row=total_row_index, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=total_row_index, column=2).font = Font(bold=True)
        ws.cell(row=total_row_index, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=total_row_index, column=2).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# to index page
@app.route("/index", methods=["GET", "POST"])
@login_required
def index():
    summary = None
    mode = "date"

    # check if user has a pending upgrade request
    pending_request = PaymentRequest.query.filter_by(user_id=current_user.id, status='pending').first()

    if request.method == "GET" and request.args.get("clear"):
        session.pop('latest_summary', None)
        session.pop('latest_mode', None)
        flash("Summary cleared.")
        return redirect(url_for("index"))

    from_date_str = request.form.get('from_date') if request.method == "POST" else None
    to_date_str = request.form.get('to_date') if request.method == "POST" else None

    try:
        from_date = parse_date_flexible(from_date_str) if from_date_str else None
        to_date = parse_date_flexible(to_date_str) if to_date_str else None
    except ValueError as e:
        flash(str(e))
        return redirect(url_for("index"))

    if request.method == "POST":
        file = request.files.get('file')
        mode = request.form.get('mode', 'date')

        if not file or not file.filename.lower().endswith('.csv'):
            flash("Only .CSV files are supported.", "warning")
            return redirect(url_for("index"))

        if current_user.plan == 'free':
            upload_count = Upload.query.filter_by(user_id=current_user.id).count()
            if upload_count >= 5:
                flash("Free plan limit reached. Upgrade to premium to upload more files.")
                return redirect(url_for('upgrade_manual'))

        filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            with open(filepath, 'rb') as f:
                summary = load_sales_data(f, mode=mode, from_date=from_date, to_date=to_date)

            if not summary:
                flash("No sales data found for the selected data range", "warning")
                return redirect(url_for("index"))

            new_upload = Upload(filename=filename, mode=mode, total=sum(summary.values()), user_id=current_user.id)
            db.session.add(new_upload)
            db.session.commit()

            session['latest_summary'] = serialize_summary(summary, mode)
            session['latest_mode'] = mode

        except Exception as e:
            flash(f"Error processing file: {e}")
            return redirect(url_for("index"))

        finally:
            if os.path.exists(filepath):
                os.remove(filepath)

    if summary is None:
        serialized = session.get('latest_summary')
        mode = session.get('latest_mode', 'date')
        if serialized:
            summary = deserialize_summary(serialized, mode)

    total_sales = sum(summary.values()) if summary else 0

    # pass pending_request to the template
    return render_template(
        "index.html",
        summary=summary,
        mode=mode,
        total_sales=total_sales,
        pending_request=pending_request
    )



# download pdf
@app.route("/download", methods=["POST"])
@login_required
def download_report():
    serialized = session.get("latest_summary")
    mode = session.get("latest_mode", "date")

    if not serialized:
        flash("No report generated yet.")
        return redirect(url_for("index"))

    summary = deserialize_summary(serialized, mode)
    report_stream = generate_excel_report(summary, mode)

    # safe download name
    filename_map = {
        "date": "sales_summary_date.xlsx",
        "item": "sales_summary_item.xlsx",
        "combined": "sales_summary_combined.xlsx"
    }
    download_name = filename_map.get(mode, "sales_summary.xlsx")

    return send_file(
        report_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

# clear form
@app.route("/clear", methods=["POST"])
@login_required
def clear_report():
    session.pop('latest_summary', None)
    session.pop('latest_mode', None)
    flash("Report cleared.")
    return redirect(url_for("index"))


# to dashbaord page
@app.route("/dashboard")
@login_required
def dashboard():
    serialized = session.get("latest_summary")
    mode = session.get("latest_mode", "date")

    if not serialized:
        flash("No report generated yet.")
        return redirect(url_for("index"))

    summary = deserialize_summary(serialized, mode)

    # prepare data for chart
    labels = []
    data = []
    if mode == "combined":
        for key, total in summary.items():
            labels.append(f"{key[0].strftime('%d/%m/%Y')} - {key[1]}")
            data.append(total)
    elif mode == "date":
        for key, total in summary.items():
            labels.append(key.strftime('%d/%m/%Y'))
            data.append(total)
    else:
        for key, total in summary.items():
            labels.append(str(key))
            data.append(total)

    return render_template("dashboard.html", labels=labels, data=data, mode=mode)



# to dashbaord_pdf page
@app.route("/download_pdf", methods=["POST"])
@login_required
def download_pdf():
    serialized = session.get("latest_summary")
    mode = session.get("latest_mode", "date")

    if not serialized:
        flash("No report generated yet.")
        return redirect(url_for("index"))

    summary = deserialize_summary(serialized, mode)

    chart_image = request.form.get("chartImage")
    chart_type = request.form.get("chartType", "bar")

    labels = []
    data = []
    for key, total in summary.items():
        if mode == "date":
            labels.append(key.strftime("%d/%m/%Y"))
        elif mode == "combined":
            labels.append(f"{key[0].strftime('%d/%m/%Y')} - {key[1]}")
        else:
            labels.append(str(key))
        data.append(total)

    summary_date = f"{labels[0]} to {labels[-1]}" if labels else "All data"
    summary_rows = list(zip(labels, data))

    rendered_HTML = render_template(
        "dashboard_pdf.html",
        labels=labels,
        data=data,
        mode=mode,
        chart_type=chart_type,
        summary_date=summary_date,
        summary_rows=summary_rows,
        chart_image=chart_image
    )

    pdf = HTML(string=rendered_HTML, base_url=request.base_url).write_pdf()

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filename=sales_chart_{mode}_{chart_type}.pdf"

    return response

# show chart ewhen exportting to pdf
@app.route("/chart_only/<type>")
def chart_only(type):
    if type not in ['bar', 'line', 'pie']:
        abort(400)  # or return a fallback view
    return render_template("chart_only.html", type=type)

# show page weher previous data were stored
@app.route("/my_uploads")
@login_required
def my_uploads():
    uploads = Upload.query.filter_by(user_id=current_user.id).order_by(Upload.uploaded_at.desc()).limit(50).all()

    return render_template("my_uploads.html", uploads=uploads)

# load previosu saved report
@app.route("/download_old_report/<int:upload_id>")
@login_required
def download_old_report(upload_id):
    upload = Upload.query.filter_by(id=upload_id, user_id=current_user.id).first_or_404()
    filename = secure_filename(upload.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if not os.path.exists(filepath):
        flash("Original file not found. Please re-upload to regenerate report.", "warning")
        return redirect(url_for("my_uploads"))

    with open(filepath, 'rb') as f:
        summary = load_sales_data(f, mode=upload.mode)

    report_stream = generate_excel_report(summary, mode=upload.mode)

    download_name = f"{upload.filename}_summary.xlsx"

    return send_file(
        report_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

# to payment page
@app.route("/upgrade_manual", methods=["GET", "POST"])
@login_required
def upgrade_manual():
    if request.method == "POST":
        proof = request.files.get('proof')
        
        if proof and proof.filename != '':
            # check file size here BEFORE saving
            proof.seek(0, os.SEEK_END)
            file_size = proof.tell()
            proof.seek(0)  # reset cursor to start
            
            MAX_UPLOAD_SIZE = 5 * 1024 * 1024  # 5 MB limit
            if file_size > MAX_UPLOAD_SIZE:
                flash("File too large. Maximum 5MB allowed.", "warning")
                return redirect(request.url)
            
            # Then check file extension
            if proof.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.pdf')):
                filename = f"{current_user.id}_{uuid.uuid4().hex}_{secure_filename(proof.filename)}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                proof.save(filepath)
                
                # save payment request to DB
                payment_request = PaymentRequest(
                    user_id=current_user.id,
                    proof_filename=filename,
                    status="pending"
                )
                db.session.add(payment_request)
                db.session.commit()

                flash("Payment proof uploaded! We'll verify and upgrade you ASAP.", "success")
                return redirect(url_for('index'))
            else:
                flash("Please upload a valid image/pdf file as proof.", "warning")
        else:
            flash("No file selected.", "warning")

    bank_info = {
        "bank_name": "Baiduri Bank",
        "account_number": "1000740511369",
        "account_name": "BATRISYIA NAJIHAH BINTI SAFRI",
        "swift_code": "BAIDBNBB"
    }
    return render_template("upgrade_manual.html", bank_info=bank_info)

# to accept payment
@app.route('/admin/payment_request/<int:request_id>/approve', methods=['POST'])
@login_required
@admin_required
def approve_payment(request_id):
    payment = PaymentRequest.query.get_or_404(request_id)
    if payment.status != 'pending':
        flash('Payment request already processed.', 'warning')
        return redirect(url_for('admin'))

    payment.status = 'approved'
    payment.user.plan = 'premium'
    db.session.commit()

    # Send approval email
    msg = Message(
        subject="Payment Approved - Your Account Upgraded",
        recipients=[payment.user.email],
        body=f"Hi {payment.user.username},\n\nYour payment has been approved and your account upgraded to premium. Enjoy!\n\nThanks,\nbat2025"
    )
    mail.send(msg)

    flash(f'User {payment.user.username} upgraded to premium!', 'success')
    return redirect(url_for('admin'))

# reject payment
@app.route('/admin/payment_request/<int:request_id>/reject', methods=['POST'])
@login_required
@admin_required
def reject_payment(request_id):
    payment = PaymentRequest.query.get_or_404(request_id)
    if payment.status != 'pending':
        flash(f'Payment request #{request_id} has already been processed.', 'warning')
        return redirect(url_for('admin'))
    payment.status = 'rejected'
    db.session.commit()

    # send rejection email
    msg = Message(
        subject="Payment Rejected",
        recipients=[payment.user.email],
        body=f"Hi {payment.user.username},\n\nUnfortunately, your payment proof was rejected. Please try again or contact support.\n\nThanks,\nbat2025"
    )
    mail.send(msg)

    flash(f'Payment request #{request_id} rejected', 'success')
    return redirect(url_for('admin'))

# to upgrade page
@app.route("/process_upgrade", methods=["POST"])
@login_required
def process_upgrade():
    # check if there's already a pending request to avoid duplicates
    existing_request = PaymentRequest.query.filter_by(user_id=current_user.id, status='pending').first()
    if existing_request:
        flash("You already have a pending upgrade request.", "warning")
        return redirect(url_for("index"))
    
    # create a new payment/upgrade request
    new_request = PaymentRequest(
        user_id=current_user.id,
        amount=3.99,  # or your premium price
        currency="BND",
        status="pending"
    )
    db.session.add(new_request)
    db.session.commit()

    flash("Upgrade request submitted! Please wait for admin approval.", "info")
    return redirect(url_for("index"))

# to payment status page
@app.route('/payment_status')
@login_required
def payment_status():
    # get all payment requests for current user, most recent first
    payments = PaymentRequest.query.filter_by(user_id=current_user.id).order_by(PaymentRequest.created_at.desc()).all()
    return render_template('payment_status.html', payments=payments)

# default page
@app.route("/")
def landing():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    return render_template("landing.html")


# terms page
@app.route("/terms")
def terms():
    return render_template("terms.html")

# sec things
@app.after_request
def add_csp_header(response):
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net;"
    )
    return response

# 404 - Page Not Found
@app.errorhandler(404)
def not_found_error(error):
    return render_template("404.html"), 404

# 400 - Bad Request
@app.errorhandler(400)
def bad_request_error(error):
    return render_template("400.html"), 400

# 500 - Internal Server Error
@app.errorhandler(500)
def internal_error(error):
    return render_template("500.html"), 500

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        create_admin_user()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True) 
    # false if in production

