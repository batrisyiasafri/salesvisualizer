from flask import Flask, render_template, request, send_file
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
import io

app = Flask(__name__)

def load_sales_data(file_stream):
    sales = defaultdict(float)
    file_stream.seek(0)
    reader = csv.DictReader(io.StringIO(file_stream.read().decode('utf-8')))
    for row in reader:
        item = row['item'].strip()
        if not item:
            continue
        amount = float(row['amount'])
        sales[item] += amount
    return sales

def generate_excel_report(summary):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Summary"

    # header row
    ws.append(["Item", "Total Sales ($)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # data rows
    for item, total in summary.items():
        ws.append([item, total])

    # format currency and center align
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        column_letter = col[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 2

    # save to BytesIO for sending as download
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

@app.route("/", methods=["GET", "POST"])
def index():
    summary = None
    error = None
    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            error = "Please upload a CSV file."
        else:
            try:
                sales = load_sales_data(file.stream)
                summary = dict(sales)
            except Exception as e:
                error = f"Error processing file: {e}"

    return render_template("index.html", summary=summary, error=error)

@app.route("/download", methods=["POST"])
def download():

    file = request.files.get("file")
    if not file or file.filename == "":
        return "No file uploaded", 400

    try:
        sales = load_sales_data(file.stream)
        summary = dict(sales)
        excel_stream = generate_excel_report(summary)
        return send_file(
            excel_stream,
            download_name="sales_summary_report.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return f"Error generating report: {e}", 500


if __name__ == "__main__":
    app.run(debug=True)
